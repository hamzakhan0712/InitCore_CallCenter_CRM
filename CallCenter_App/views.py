import json
import csv
import io
import os
import re
import pandas as pd
import uuid
from uuid import uuid4
from io import BytesIO
from datetime import datetime, timedelta, date
import openpyxl
import pdfkit
from num2words import num2words
from django.conf import settings
from django.db import IntegrityError
from dateutil import parser as date_parser
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.core.management import call_command
from django.db import transaction
from django.db.models import Sum, Q, Count
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.defaultfilters import floatformat
from django.template.loader import render_to_string
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth.models import User
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .forms import (
    CustomUserCreationForm, CustomUserChangeForm, CustomAuthenticationForm,
    CompanyForm,ComplaintForm, LeadImportForm, PaidCustomerForm,
    PaymentMethodForm, TeamForm, AddAgentToTeamForm, LeadForm, BreakTypeForm,
    PackageForm, SubDispositionForm, UpdateSalesForm, PaymentMethod, Package
)
from .models import (
    Team, Attendance, BreakType, Break, UserProfile, Complaint, Lead,
    LeadTransferRecord, SubDisposition, PaidCustomer, Company, Invoice,
    InvoicePDF, AgentSalesHistory
)
from .utils import record_action, record_agent_sales_history

##############################################################################################################################################

@login_required
def navbar(request):
    break_types = BreakType.objects.all()
    
    context = {
        'break_types': break_types,
    }
    return render(request, 'navbar .html', context)

def is_superuser(user):
    return user.is_superuser

############CLASS######CLASS#####CLASS###CLASS#####CLASS#####CLASS#####CLASS#######CLASSCLASS#######CLASS##################################################################################

class CustomLoginView(LoginView):
    template_name = 'login.html'
    authentication_form = CustomAuthenticationForm

    def form_valid(self, form):
        response = super().form_valid(form)
        if not self.request.user.is_superuser:
            self.mark_attendance(self.request.user.profile, 'login') 
        messages.success(self.request, 'You have been successfully logged in.')
        return response

    def mark_attendance(self, user, action):
        today = timezone.localtime(timezone.now()).date()
        attendance, created = Attendance.objects.get_or_create(user=user, date=today)
        if action == 'login':
            now = timezone.localtime(timezone.now())
            if not attendance.login_time:
                attendance.login_time = now.time()
            attendance.last_login_time = now.time()
            attendance.save()

class CustomLogoutView(LogoutView):
    next_page = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            self.mark_attendance(request.user.profile, 'logout') 
        return super().dispatch(request, *args, **kwargs)

    def mark_attendance(self, user, action):
        today = timezone.localtime(timezone.now()).date()
        attendance = Attendance.objects.filter(user=user, date=today).last()
        if attendance:
            now = timezone.localtime(timezone.now())
            if action == 'logout':
                attendance.logout_time = now.time()
                attendance.last_logout_time = now.time()
                attendance.save()

@login_required
def monitor_users(request):
    return render(request, 'monitor_users.html')

@login_required
def get_recent_breaks(request):
    if not request.user.is_authenticated or not (request.user.is_superuser or request.user.profile.role == 'Team Leader'):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.user.is_superuser:
        recent_breaks = Break.objects.filter(
            start_time__gte=timezone.localtime(timezone.now()) - timezone.timedelta(days=1)
        ).select_related('user', 'break_type').values(
            'user_id', 'user__user__first_name', 'user__user__last_name', 'user__role', 'break_type__name', 'start_time', 'active'
        )
    elif request.user.profile.role == 'Team Leader':
        user_team = Team.objects.filter(leader=request.user.profile).first()
        if user_team:
            team_users = user_team.agents.all()
            recent_breaks = Break.objects.filter(
                start_time__gte=timezone.localtime(timezone.now()) - timezone.timedelta(days=1),
                user__in=team_users
            ).select_related('user', 'break_type').values(
                'user_id', 'user__user__first_name', 'user__user__last_name', 'user__role', 'break_type__name', 'start_time', 'active'
            )
        else:
            recent_breaks = []

    data = [{
        'user_id': b['user_id'],
        'user_name': f"{b['user__user__first_name']} {b['user__user__last_name']}",
        'user_role': b['user__role'],
        'break_type': b['break_type__name'],
        'start_time': timezone.localtime(b['start_time']).strftime('%Y-%m-%d %H:%M:%S'),
        'on_break': b['active'],
    } for b in recent_breaks]

    return JsonResponse(data, safe=False)

@login_required
def break_state(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)
        user_break = Break.objects.filter(user=user, active=True).first()
        if user_break:
            response = {
                'success': True,
                'message': 'Break state retrieved successfully.',
                'on_break': True,
                'start_time': user_break.start_time.isoformat(),
                'break_type_id': user_break.break_type.id,
            }
        else:
            response = {
                'success': True,
                'message': 'User is not on break.',
                'on_break': False,
                'start_time': None,
                'break_type_id': None,
            }
    except User.DoesNotExist:
        response = {
            'success': False,
            'error': 'User does not exist.',
        }
    except Exception as e:
        response = {
            'success': False,
            'error': f'An error occurred: {str(e)}',
        }
    return JsonResponse(response)


@login_required
def dashboard(request):
    user_profile = request.user.profile
    today = timezone.localtime(timezone.now()).date()
    start_of_month = today.replace(day=1)

    sales_data = []
    labels = []
    all_paid_customers = PaidCustomer.objects.none()  
    all_leads = Lead.objects.none()  
    attendance_rate = 0
    total_present = 0
    total_absent = 0
    total_half_day = 0

    if request.user.is_superuser:
        amount_paid = Invoice.objects.filter(customer__payment_status='completed').aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
        unique_customers_today = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date=today).values('contact_number').distinct().count()
        today_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__payment_date=today).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
        unique_customers_this_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__gte=start_of_month).values('contact_number').distinct().count()
        this_month_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__payment_date__gte=start_of_month).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_end = today.replace(day=1) - timedelta(days=1)

        unique_customers_last_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__range=[last_month_start, last_month_end]).values('contact_number').distinct().count()

        teams = Team.objects.all()
        team_leader_sales_last_month = {}
        for team in teams:
            total_sales_last_month = Invoice.objects.filter(
                customer__lead__assigned_to_team=team,
                customer__payment_status='completed',
                customer__verified=True,
                customer__payment_date__range=[last_month_start, last_month_end]
            ).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
            team_leader_name = f"{team.leader.user.first_name} {team.leader.user.last_name}"
            team_leader_sales_last_month[team_leader_name] = total_sales_last_month

        team_leader_sales = {}
        for team in teams:
            total_sales_amount = Invoice.objects.filter(
                customer__lead__assigned_to_team=team,
                customer__payment_status='completed',
                customer__verified=True
            ).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
            team_leader_name = f"{team.leader.user.first_name} {team.leader.user.last_name}"
            team_leader_sales[team_leader_name] = total_sales_amount

        total_attendances = Attendance.objects.count()
        total_present = Attendance.objects.filter(status='Present').count()
        total_absent = Attendance.objects.filter(status='Absent').count()
        total_half_day = Attendance.objects.filter(status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100

        thirty_days_ago = today - timedelta(days=30)
        current_date = thirty_days_ago
        while current_date <= today:
            sales_amount = Invoice.objects.filter(
                customer__payment_status='completed',
                customer__verified=True,
                customer__payment_date=current_date
            ).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

            sales_data.append(float(sales_amount))
            labels.append(current_date.strftime('%Y-%m-%d'))

            current_date += timedelta(days=1)

        all_paid_customers = PaidCustomer.unique_paid_customers()
        all_leads = Lead.objects.all()

    elif user_profile.role == 'Team Leader':
        teams = Team.objects.filter(leader=user_profile)
        team_agents = UserProfile.objects.filter(teams_as_agent__in=teams)

        amount_paid = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__lead__assigned_to__in=team_agents).aggregate(total_amount_paid=Sum('customer__amount_with_gst'))['total_amount_paid'] or 0
        unique_customers_today = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date=today, lead__assigned_to__in=team_agents).values('contact_number').distinct().count()
        today_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__payment_date=today, customer__lead__assigned_to__in=team_agents).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
        unique_customers_this_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__gte=start_of_month, lead__assigned_to__in=team_agents).values('contact_number').distinct().count()
        this_month_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__payment_date__gte=start_of_month, customer__lead__assigned_to__in=team_agents).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_end = today.replace(day=1) - timedelta(days=1)
        unique_customers_last_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__range=[last_month_start, last_month_end], lead__assigned_to__in=team_agents).values('contact_number').distinct().count()

        team_leader_sales_last_month = {user_profile.user.get_full_name(): unique_customers_last_month}
        team_leader_sales = {user_profile.user.get_full_name(): this_month_sales_amount}

        total_attendances = Attendance.objects.filter(user__in=team_agents).count()
        total_present = Attendance.objects.filter(user__in=team_agents, status='Present').count()
        total_absent = Attendance.objects.filter(user__in=team_agents, status='Absent').count()
        total_half_day = Attendance.objects.filter(user__in=team_agents, status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100

        thirty_days_ago = today - timedelta(days=30)
        current_date = thirty_days_ago
        while current_date <= today:
            sales_amount = Invoice.objects.filter(
                customer__payment_status='completed',
                customer__verified=True,
                customer__payment_date=current_date,
                customer__lead__assigned_to__in=team_agents
            ).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

            sales_data.append(float(sales_amount))
            labels.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        if teams.exists(): 
            team = teams.first()  
            all_paid_customers = PaidCustomer.objects.filter(lead__assigned_to_team=team)
            all_leads = Lead.objects.filter(assigned_to_team=team)

    elif user_profile.role == 'Agent':
        amount_paid = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__lead__assigned_to=user_profile).aggregate(total_amount_paid=Sum('customer__amount_with_gst'))['total_amount_paid'] or 0
        unique_customers_today = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date=today, lead__assigned_to=user_profile).values('contact_number').distinct().count()
        today_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__payment_date=today, customer__lead__assigned_to=user_profile).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
        unique_customers_this_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__gte=start_of_month, lead__assigned_to=user_profile).values('contact_number').distinct().count()
        this_month_sales_amount = Invoice.objects.filter(customer__payment_status='completed', customer__verified=True, customer__payment_date__gte=start_of_month, customer__lead__assigned_to=user_profile).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_end = today.replace(day=1) - timedelta(days=1)
        unique_customers_last_month = PaidCustomer.objects.filter(payment_status='completed', verified=True, payment_date__range=[last_month_start, last_month_end], lead__assigned_to=user_profile).values('contact_number').distinct().count()

        team_leader_sales_last_month = {user_profile.user.get_full_name(): unique_customers_last_month}
        team_leader_sales = {user_profile.user.get_full_name(): this_month_sales_amount}

        total_attendances = Attendance.objects.filter(user=user_profile).count()
        total_present = Attendance.objects.filter(user=user_profile, status='Present').count()
        total_absent = Attendance.objects.filter(user=user_profile, status='Absent').count()
        total_half_day = Attendance.objects.filter(user=user_profile, status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100

        thirty_days_ago = today - timedelta(days=30)
        current_date = thirty_days_ago
        while current_date <= today:
            sales_amount = Invoice.objects.filter(
                customer__payment_status='completed',
                customer__verified=True,
                customer__payment_date=current_date,
                customer__lead__assigned_to=user_profile
            ).aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0

            sales_data.append(float(sales_amount))
            labels.append(current_date.strftime('%Y-%m-%d'))

            current_date += timedelta(days=1)

        all_paid_customers = PaidCustomer.objects.filter(lead__assigned_to=user_profile)
        all_leads = Lead.objects.filter(assigned_to=user_profile)

    paid_customers_overall_look = {
        'labels': ['Pending', 'Completed', 'Failed'],
        'data': [
            all_paid_customers.filter(payment_status='pending').count(),
            all_paid_customers.filter(payment_status='completed').count(),
            all_paid_customers.filter(payment_status='failed').count()
        ]
    }

    all_leads_overall_look = {
        'labels': ['Fresh', 'Connected', 'Not connected'],
        'data': [
            all_leads.filter(disposition='Fresh').count(),
            all_leads.filter(disposition='Connected').count(),
            all_leads.filter(disposition='Not connected').count()
        ]
    }

    attendance_overall_look = {
        'labels': ['Present', 'Absent', 'Half day'],
        'data': [total_present, total_absent, total_half_day]
    }

    context = {
        'sales_data': sales_data,
        'sales_labels': labels,
        'all_leads': all_leads,
        'attendance_rate': attendance_rate,
        'all_paid_customers': all_paid_customers,
        'team_leader_sales': team_leader_sales,
        'amount_paid': amount_paid,
        'unique_customers_today': unique_customers_today,
        'today_sales_amount': today_sales_amount,
        'unique_customers_this_month': unique_customers_this_month,
        'this_month_sales_amount': this_month_sales_amount,
        'unique_customers_last_month': unique_customers_last_month,
        'team_leader_sales_last_month': team_leader_sales_last_month,
        'paid_customers_overall_look': paid_customers_overall_look,
        'all_leads_overall_look': all_leads_overall_look,
        'attendance_overall_look': attendance_overall_look,
    }

    return render(request, 'dashboard.html', context)


############STAFF########STAFF#########STAFF######STAFF#########STAFF#######STAFF#######STAFF#########STAFF###########################################################################

@login_required
def staff_list(request):
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    page = request.GET.get('page', 1)

    users_profile = UserProfile.objects.all()

    if search_query:
        users_profile = users_profile.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(address__icontains=search_query)
        )

    if start_date:
        users_profile = users_profile.filter(date_of_joining__gte=start_date)
    if end_date:
        users_profile = users_profile.filter(date_of_joining__lte=end_date)

    if sort_by == 'full_name':
        users_profile = users_profile.order_by('user__first_name', 'user__last_name')
    elif sort_by == 'date_of_joining':
        users_profile = users_profile.order_by('date_of_joining')
    elif sort_by == 'status':
        users_profile = users_profile.order_by('status')
    else:
        users_profile = users_profile.order_by('user__first_name', 'user__last_name')

    team_leaders = users_profile.filter(role='Team Leader')
    agents = users_profile.filter(role='Agent')
    teams = Team.objects.all()
    my_team = None
    my_team_members = UserProfile.objects.none()

    if request.user.profile.role == 'Team Leader':
        my_team = Team.objects.filter(leader=request.user.profile).first()
        if my_team:
            my_team_members = my_team.agents.select_related('user')

            if search_query:
                my_team_members = my_team_members.filter(
                    Q(user__username__icontains=search_query) |
                    Q(user__email__icontains=search_query) |
                    Q(phone_number__icontains=search_query) |
                    Q(user__first_name__icontains=search_query) |
                    Q(user__last_name__icontains=search_query) |
                    Q(address__icontains=search_query)
                )

            if start_date:
                my_team_members = my_team_members.filter(date_of_joining__gte=start_date)
            if end_date:
                my_team_members = my_team_members.filter(date_of_joining__lte=end_date)

            if sort_by == 'full_name':
                my_team_members = my_team_members.order_by('user__first_name', 'user__last_name')
            elif sort_by == 'date_of_joining':
                my_team_members = my_team_members.order_by('date_of_joining')
            elif sort_by == 'status':
                my_team_members = my_team_members.order_by('status')
            else:
                my_team_members = my_team_members.order_by('user__first_name', 'user__last_name')

    team_leaders_paginator = Paginator(team_leaders, 10)  
    agents_paginator = Paginator(agents, 10) 
    my_team_members_paginator = Paginator(my_team_members, 10)  

    try:
        team_leaders_page = team_leaders_paginator.page(page)
    except PageNotAnInteger:
        team_leaders_page = team_leaders_paginator.page(1)
    except EmptyPage:
        team_leaders_page = team_leaders_paginator.page(team_leaders_paginator.num_pages)

    try:
        agents_page = agents_paginator.page(page)
    except PageNotAnInteger:
        agents_page = agents_paginator.page(1)
    except EmptyPage:
        agents_page = agents_paginator.page(agents_paginator.num_pages)

    try:
        my_team_members_page = my_team_members_paginator.page(page)
    except PageNotAnInteger:
        my_team_members_page = my_team_members_paginator.page(1)
    except EmptyPage:
        my_team_members_page = my_team_members_paginator.page(my_team_members_paginator.num_pages)

    context = {
        'search_query': search_query,
        'sort_by': sort_by,
        'start_date': start_date,
        'end_date': end_date,
        'team_leaders': team_leaders_page,
        'agents': agents_page,
        'teams': teams,
        'my_team': my_team,
        'my_team_members': my_team_members_page,
        'is_superuser': request.user.is_superuser,
    }
    return render(request, 'staff_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'User created successfully.')
            return redirect('create_user')
        else:
            messages.error(request, 'Form submission failed. Please check the form for errors.')
    else:
        form = CustomUserCreationForm(request=request)
    return render(request, 'create_user.html', {'form': form})


@login_required
def edit_user(request, user_id):
    user_profile = get_object_or_404(UserProfile, id=user_id)
    user = user_profile.user

    if request.method == 'POST':
        user_form = CustomUserChangeForm(request.POST, instance=user)
        if user_form.is_valid():
            user = user_form.save()
            messages.success(request, 'User updated successfully.')
        else:
            messages.error(request, 'Form submission failed. Please check the form for errors.')
    else:
        user_form = CustomUserChangeForm(instance=user)

    return render(request, 'edit_user.html', {'form': user_form})


@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user.delete()
        return redirect('staff_list')

    context = {
        'user': user
    }
    return render(request, 'delete_user_confirm.html', context)

@csrf_exempt
@require_POST
def update_user_status(request, user_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in dict(UserProfile.STATUS_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid status value.'}, status=400)
        
        user_profile = get_object_or_404(UserProfile, id=user_id)
        user_profile.status = new_status
        user_profile.save()
        
        return JsonResponse({'success': True})
    except (ValueError, KeyError, json.JSONDecodeError):
        return JsonResponse({'success': False, 'error': 'Invalid request data.'}, status=400)
    
##########TEAMS##########TEAMS####TEAMS#######TEAMS#######TEAMS############TEAMS########TEAMS#####TEAMS####TEAMS###########################################################################

@login_required
def create_team(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')

    if request.method == 'POST':
        form = TeamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Team created successfully.')
            return redirect('staff_list')
        else:
            messages.error(request, 'There was an error creating the team. Please check the form and try again.')
    else:
        form = TeamForm()
        team_leaders = UserProfile.objects.filter(role='Team Leader')
        form.fields['leader'].queryset = team_leaders  

    return render(request, 'create_team.html', {'form': form})

@login_required
def edit_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    if request.method == 'POST':
        form = TeamForm(request.POST, instance=team)
        if form.is_valid():
            form.save()
            return redirect('staff_list')
    else:
        form = TeamForm(instance=team)
    return render(request, 'edit_team.html', {'form': form})

@login_required
def delete_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    if request.method == 'POST':
        team.delete()
        return redirect('staff_list')
    return render(request, 'delete_team.html', {'team': team})

def add_agent_to_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    message = None
    if request.method == 'POST':
        form = AddAgentToTeamForm(request.POST)
        if form.is_valid():
            agent = form.cleaned_data['agent']
            team.agents.add(agent)
            message = 'Agent added to the team successfully.'
    else:
        form = AddAgentToTeamForm()
    return JsonResponse({'message': message})

@login_required
def remove_agent_from_team(request, team_id, agent_id):
    team = get_object_or_404(Team, id=team_id)
    agent = get_object_or_404(User, id=agent_id)
    agent_profile = agent.profile

    if request.method == 'POST':
        if agent_profile in team.agents.all():
            team.agents.remove(agent_profile)
            messages.success(request, 'Agent has been successfully removed from the team.')
        else:
            messages.error(request, 'Agent is not a member of this team.')

        return redirect('staff_list')

    context = {
        'team': team,
        'agent': agent
    }
    return render(request, 'remove_agent_confirm.html', context)

##########LEADS###########LEADS################LEADS#####LEADS#####LEADS####LEADS#####LEADS#####LEADS#######LEADS############################################################################


@login_required
def lead_list(request):
    search_query = request.GET.get('search', '')
    disposition = request.GET.get('disposition', '')
    sub_disposition = request.GET.get('sub_disposition', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sort_by = request.GET.get('sort', 'id')
    page = request.GET.get('page', 1)

    leads = Lead.objects.all()
    teams = Team.objects.all()
    agents = UserProfile.objects.filter(role='Agent')
    other_teams = Team.objects.none()
    team_members = []
    my_team_members = []

    if request.user.profile.role == 'Team Leader':
        team = Team.objects.filter(leader=request.user.profile).first()
        if team:
            team_members = team.agents.all()
            leads = leads.filter(Q(assigned_to__in=team_members) | Q(assigned_to_team=team))
            other_teams = Team.objects.all().exclude(leader=request.user.profile)
        else:
            leads = Lead.objects.none()
            other_teams = Team.objects.all().exclude(leader=request.user.profile)

    elif request.user.profile.role == 'Agent':
        leads = leads.filter(assigned_to=request.user.profile)
        my_team = Team.objects.filter(agents=request.user.profile).first()
        if my_team:
            my_team_members = my_team.agents.exclude(id=request.user.profile.id).select_related('user')

    if disposition:
        leads = leads.filter(disposition=disposition)

    if sub_disposition:
        leads = leads.filter(sub_disposition__name__icontains=sub_disposition)

    if start_date:
        leads = leads.filter(date__gte=start_date)

    if end_date:
        leads = leads.filter(date__lte=end_date)

    leads = leads.order_by(sort_by)

    paginator = Paginator(leads, 10)  

    try:
        leads = paginator.page(page)
    except PageNotAnInteger:
        leads = paginator.page(1)
    except EmptyPage:
        leads = paginator.page(paginator.num_pages)

    if request.method == 'POST' and 'file' in request.FILES:
        form = LeadImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            file_extension = uploaded_file.name.split('.')[-1].lower()

            if file_extension == 'csv':
                file_data = uploaded_file.read().decode('utf-8')
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file)
                file_data = df.to_csv(index=False)
            else:
                messages.error(request, 'Unsupported file format. Please upload a CSV or XLSX file.')
                return redirect('lead_list')

            request.session['uploaded_file'] = file_data
            return redirect('lead_mapping')
    else:
        form = LeadImportForm()

    if search_query:
        leads = Lead.objects.filter(
            Q(full_name__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )

    context = {
        'leads': leads,
        'teams': teams,
        'agents': agents,
        'search_query': search_query,
        'disposition': disposition,
        'sub_disposition': sub_disposition,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'lead_disposition_choices': Lead.DISPOSITION_CHOICES,
        'sub_disposition_choices': SubDisposition.objects.all(),
        'is_superuser': request.user.is_superuser,
        'form': form,
        'team_members': team_members,
        'my_team_members': my_team_members,
        'other_teams': other_teams,
    }

    return render(request, 'lead_list.html', context)

@login_required
def lead_mapping(request):
    file_data = request.session.get('uploaded_file')
    if not file_data:
        messages.error(request, 'No file data found. Please upload a file first.')
        return redirect('lead_list')

    csv_reader = csv.reader(io.StringIO(file_data))
    header = next(csv_reader)
    lead_fields = [field.name.lower() for field in Lead._meta.get_fields() if field.concrete]
    date_fields = ['date', 'reminder']

    if request.method == 'POST':
        mapping = request.POST.getlist('mapping')
        for row in csv_reader:
            lead_data = {}
            for i in range(len(lead_fields)):
                if mapping[i]:
                    field_name = lead_fields[i]
                    try:
                        column_index = header.index(next(col for col in header if col.lower() == mapping[i].lower()))
                        value = row[column_index]
                    except StopIteration:
                        messages.error(request, f'No matching column found for field {field_name}: {mapping[i]}')
                        return redirect('lead_mapping')
                    
                    if field_name in date_fields:
                        try:
                            parsed_date = date_parser.parse(value)
                            value = parsed_date.strftime('%Y-%m-%d')
                        except (ValueError, date_parser.ParserError):
                            messages.error(request, f'Invalid date format for field {field_name}: {value}. Please ensure it is a valid date.')
                            return redirect('lead_mapping')

                    lead_data[field_name] = value

            contact_number = lead_data.get('contact_number')
            if contact_number:
                try:
                    Lead.objects.create(**lead_data)
                except IntegrityError:
                    Lead.objects.filter(contact_number=contact_number).update(**lead_data)

        messages.success(request, 'Leads imported successfully.')
        return redirect('lead_list')

    context = {
        'header': header,
        'lead_fields': lead_fields,
    }
    return render(request, 'lead_mapping.html', context)

def lead_history(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    history = lead.history.all().order_by('-timestamp')
    return render(request, 'lead_history.html', {'lead': lead, 'history': history})

@login_required
def create_lead(request):
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)

            if request.user.is_superuser:
                pass
            elif request.user.profile.role == 'Team Leader':
                lead.assigned_to_team = request.user.teams_as_leader.first()
            elif request.user.profile.role == 'Agent':
                lead.assigned_to_team = request.user.teams_as_agent.first()
                lead.assigned_to = request.user

            lead.save()
            record_action(lead, 'Lead Created', request.user.username)
            return redirect('lead_list')
    else:
        form = LeadForm()

    return render(request, 'create_lead.html', {'form': form})

def export_leads(request):
    leads = Lead.objects.all()
    current_date = datetime.now().strftime('%Y-%m-%d')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="lead_export_{current_date}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Lead ID', 'Date', 'Full Name', 'Contact Number', 'State', 'Capital', 'Assigned To', 'Assigned to Team', 'Disposition', 'Sub Disposition', 'Remark'])
    
    for lead in leads:
        writer.writerow([
            lead.id,
            lead.date,
            lead.full_name,
            lead.contact_number,
            lead.state,
            lead.capital,
            lead.assigned_to,
            lead.assigned_to_team,
            lead.disposition,
            lead.sub_disposition.name if lead.sub_disposition else '',  
            lead.remark
        ])
    
    record_action(None, 'Leads Exported', request.user.username, 'Exported all leads to CSV')
    return response

def edit_lead(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            form.save()
            record_action(lead, 'Lead Updated', request.user.username)
            return redirect('lead_list')
    else:
        form = LeadForm(instance=lead)
    return render(request, 'edit_lead.html', {'form': form})

@login_required
@require_POST
def assign_leads_to_team(request):
    user_profile = request.user.profile
    selected_leads = request.POST.getlist('selected_leads')
    transfer_note = request.POST.get('transfer_note', '')

    if not selected_leads:
        messages.error(request, 'No leads selected.')
        return redirect('lead_list')

    leads = Lead.objects.filter(id__in=selected_leads)

    if request.user.is_superuser:
        team_id = request.POST.get('team')
        agent_id = request.POST.get('agent')
        
        if team_id:
            team = Team.objects.get(id=team_id)
            leads.update(assigned_to_team=team)
            for lead in leads:
                record_action(lead, 'Lead Assigned', user_profile.user.username, f'Assigned to team {team.name}', transfer_note)
            messages.success(request, 'Leads assigned to team successfully.')
        
        elif agent_id:
            agent_profile = UserProfile.objects.get(user_id=agent_id)
            for lead in leads:
                lead.assigned_to = agent_profile
                lead.save()
                record_action(lead, 'Lead Assigned', user_profile.user.username, f'Assigned to agent {agent_profile.user.get_full_name()}', transfer_note)
            messages.success(request, 'Leads assigned to agent successfully.')

    elif user_profile.role == 'Team Leader':
        team_member_id = request.POST.get('team_member')
        new_team_id = request.POST.get('other_team_leader')
        if team_member_id:
            team_member_profile = UserProfile.objects.get(user_id=team_member_id)
            for lead in leads:
                lead.assigned_to = team_member_profile.user.profile
                lead.save()
                record_action(lead, 'Lead Assigned', user_profile.user.get_full_name(), f'Assigned to team member {team_member_profile.user.get_full_name()}', transfer_note)
            messages.success(request, 'Leads assigned to team member successfully.')
        
        elif new_team_id:
            new_team = Team.objects.get(id=new_team_id)
            leads.update(assigned_to=None)  
            leads.update(assigned_to_team=new_team)
            for lead in leads:
                LeadTransferRecord.objects.create(
                    lead=lead,
                    from_user=user_profile,
                    to_user=new_team.leader,
                    transfer_remark=transfer_note,
                    disposition=lead.disposition,
                    sub_disposition=lead.sub_disposition
                )
                record_action(lead, 'Lead Transferred', user_profile.user.get_full_name(), f'Transferred to team {new_team.name}', transfer_note)
            messages.success(request, 'Leads transferred to new team successfully.')

    elif user_profile.role == 'Agent':
        other_agent_id = request.POST.get('other_agent')
        
        if other_agent_id:
            other_agent_profile = UserProfile.objects.get(user_id=other_agent_id)
            for lead in leads:
                lead.assigned_to = other_agent_profile.user.profile
                lead.save()
                LeadTransferRecord.objects.create(
                    lead=lead,
                    from_user=user_profile,
                    to_user=other_agent_profile,
                    transfer_remark=transfer_note,
                    disposition=lead.disposition,
                    sub_disposition=lead.sub_disposition
                )
                record_action(lead, 'Lead Transferred', user_profile.user.get_full_name(), f'Transferred to agent {other_agent_profile.user.get_full_name()}', transfer_note)
            messages.success(request, 'Leads transferred to other agent successfully.')

    return redirect('lead_list')


@require_POST
def dispose_lead(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        disposition = data.get('disposition')
        sub_disposition_name = data.get('sub_disposition')
        remark = data.get('remark')
        reminder_days = int(data.get('reminder_days', 0))
        reminder_hours = int(data.get('reminder_hours', 0))
        reminder_minutes = int(data.get('reminder_minutes', 0))
        sub_disposition, _ = SubDisposition.objects.get_or_create(name=sub_disposition_name)
        reminder_delta = timedelta(days=reminder_days, hours=reminder_hours, minutes=reminder_minutes)
        reminder_time = timezone.localtime(timezone.now()) + reminder_delta if reminder_delta.total_seconds() > 0 else None

        lead = Lead.objects.get(pk=lead_id)
        lead.disposition = disposition
        lead.sub_disposition = sub_disposition
        lead.remark = remark
        lead.reminder = reminder_time
        lead.save()

        record_action(lead, 'Lead Disposed', request.user.username, f'Disposition: {disposition}, Sub Disposition: {sub_disposition.name}, Remark: {remark}')
        return JsonResponse({'success': True})

    return JsonResponse({'success': False})

@login_required
def delete_lead(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    if request.method == 'POST':
        lead.delete()
        return redirect('lead_list')
    return render(request, 'delete_lead.html', {'lead': lead})


@login_required
def lead_transfers(request):
    user = request.user
    profile = user.profile
    search_query = request.GET.get('search', '')
    contact_number = request.GET.get('contact_number', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_disposition = request.GET.get('sub_disposition', '')
    sort = request.GET.get('sort', '-id')
    page = request.GET.get('page', 1)

    lead_transfers = LeadTransferRecord.objects.all()

    if contact_number:
        lead_transfers = lead_transfers.filter(lead__contact_number__icontains=contact_number)

    if start_date:
        lead_transfers = lead_transfers.filter(transfer_date__gte=start_date)

    if end_date:
        lead_transfers = lead_transfers.filter(transfer_date__lte=end_date)

    if sub_disposition:
        lead_transfers = lead_transfers.filter(sub_disposition__name__icontains=sub_disposition)

    if sort:
        lead_transfers = lead_transfers.order_by(sort)

    team_leader_transfers = LeadTransferRecord.objects.none()
    agent_transfers = LeadTransferRecord.objects.none()

    if profile.role == 'Team Leader':
        team = Team.objects.filter(leader=profile).first()
        team_leader_transfers = lead_transfers.filter(Q(from_user=user.profile) | Q(to_user=user.profile))
        if search_query:
            team_leader_transfers = team_leader_transfers.filter(
                Q(lead__full_name__icontains=search_query) |
                Q(lead__contact_number__icontains=search_query)
            )
        if team:
            agent_transfers = lead_transfers.filter(Q(from_user__in=team.agents.all()) | Q(to_user__in=team.agents.all()))
            if search_query:
                agent_transfers = agent_transfers.filter(
                    Q(lead__full_name__icontains=search_query) |
                    Q(lead__contact_number__icontains=search_query)
                )
        else:
            agent_transfers = LeadTransferRecord.objects.none()

    elif profile.role == 'Agent':
        agent_transfers = lead_transfers.filter(Q(from_user=user.profile) | Q(to_user=user.profile))
        if search_query:
            agent_transfers = agent_transfers.filter(
                Q(lead__full_name__icontains=search_query) |
                Q(lead__contact_number__icontains=search_query)
            )
        team_leader_transfers = LeadTransferRecord.objects.none()

    elif user.is_superuser:
        team_leader_transfers = lead_transfers.filter(Q(from_user__role='Team Leader') | Q(to_user__role='Team Leader'))
        agent_transfers = lead_transfers.filter(Q(from_user__role='Agent') | Q(to_user__role='Agent'))

    paginator_team_leader_transfers = Paginator(team_leader_transfers, 10)
    try:
        team_leader_transfers_page = paginator_team_leader_transfers.page(page)
    except PageNotAnInteger:
        team_leader_transfers_page = paginator_team_leader_transfers.page(1)
    except EmptyPage:
        team_leader_transfers_page = paginator_team_leader_transfers.page(paginator_team_leader_transfers.num_pages)

    paginator_agent_transfers = Paginator(agent_transfers, 10)
    try:
        agent_transfers_page = paginator_agent_transfers.page(page)
    except PageNotAnInteger:
        agent_transfers_page = paginator_agent_transfers.page(1)
    except EmptyPage:
        agent_transfers_page = paginator_agent_transfers.page(paginator_agent_transfers.num_pages)

    context = {
        'lead_transfers': lead_transfers,
        'team_leader_transfers': team_leader_transfers_page,
        'agent_transfers': agent_transfers_page,
        'sub_disposition_choices': SubDisposition.objects.all(),
        'teams': Team.objects.all(),
        'agents': UserProfile.objects.filter(role='Agent'),
    }

    return render(request, 'lead_transfers.html', context)


@login_required
def delete_lead_transfer(request, lead_id):
    lead_transfer = get_object_or_404(LeadTransferRecord, id=lead_id)
    if request.user.is_superuser:
        lead_transfer.delete()
        messages.success(request, 'Lead transfer record deleted successfully.')
    else:
        messages.error(request, 'You do not have permission to delete this lead transfer record.')
    return redirect('lead_transfers')


@login_required
def download_excel_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    team_id = request.GET.get('team')
    agent_id = request.GET.get('agent')

    lead_transfers = LeadTransferRecord.objects.all()

    if start_date:
        lead_transfers = lead_transfers.filter(transfer_date__gte=start_date)
    
    if end_date:
        lead_transfers = lead_transfers.filter(transfer_date__lte=end_date)
    
    if team_id:
        team = Team.objects.get(id=team_id)
        lead_transfers = lead_transfers.filter(Q(lead__assigned_to_team=team) | Q(from_user__teams=team) | Q(to_user__teams=team))

    if agent_id:
        agent = User.objects.get(id=agent_id)
        lead_transfers = lead_transfers.filter(Q(from_user=agent) | Q(to_user=agent))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="lead_transfers.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Contact Number', 'Full Name', 'Transfer Date', 'Transfer Time', 'From Agent', 'To Agent', 'Transfer Remark', 'Disposition', 'Sub Disposition', 'Lead Remark'])

    for transfer in lead_transfers:
        writer.writerow([
            transfer.id,
            transfer.lead.contact_number,
            transfer.lead.full_name,
            transfer.transfer_date,
            transfer.transfer_time,
            transfer.from_user.get_full_name(),
            transfer.to_user.get_full_name() if transfer.to_user else 'N/A',
            transfer.transfer_remark,
            transfer.disposition,
            transfer.sub_disposition.name,
            transfer.lead.remark,
        ])

    return response

########CUSTOMERS######CUSTOMERS#####CUSTOMERS#########CUSTOMERS######CUSTOMERS######CUSTOMERS#######CUSTOMERS####################################################################################################


@login_required
def paid_customers(request):
    user = request.user

    search_query = request.GET.get('search', '')
    disposition = request.GET.get('disposition', '')
    sub_disposition = request.GET.get('sub_disposition', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sort_by = request.GET.get('sort', 'id')
    payment_method = request.GET.get('payment_method', '')
    payment_status = request.GET.get('payment_status', '')
    team_leader_id = request.GET.get('team_leader', '')
    page = request.GET.get('page', 1)

    if user.is_superuser:
        paid_customers = PaidCustomer.objects.all()
    elif user.profile.role == 'Team Leader':
        team = Team.objects.filter(leader=user.profile).first()
        if team:
            team_members = team.agents.all()
            paid_customers = PaidCustomer.objects.filter(lead__assigned_to__in=team_members)
        else:
            paid_customers = PaidCustomer.objects.none()
    elif user.profile.role == 'Agent':
        paid_customers = PaidCustomer.objects.filter(lead__assigned_to=user.profile)
    else:
        paid_customers = PaidCustomer.objects.none()

    if search_query:
        paid_customers = paid_customers.filter(
            Q(lead__full_name__icontains=search_query) | 
            Q(lead__contact_number__icontains=search_query) |
            Q(customer_id__icontains=search_query) 
        )

    if disposition:
        paid_customers = paid_customers.filter(lead__disposition=disposition)

    if sub_disposition:
        paid_customers = paid_customers.filter(lead__sub_disposition__name__icontains=sub_disposition)

    if start_date:
        paid_customers = paid_customers.filter(payment_date__gte=start_date)

    if end_date:
        paid_customers = paid_customers.filter(payment_date__lte=end_date)

    if payment_method:
        paid_customers = paid_customers.filter(payment_method__id=payment_method)

    if payment_status:
        paid_customers = paid_customers.filter(payment_status=payment_status)

    if team_leader_id and user.is_superuser:
        team_leader = UserProfile.objects.filter(id=team_leader_id).first()
        if team_leader:
            team_members = Team.objects.filter(leader=team_leader).first().agents.all()
            paid_customers = paid_customers.filter(lead__assigned_to__in=team_members)

    paid_customers = paid_customers.order_by(sort_by)
    customer_invoices = []
    for customer in paid_customers:
        invoice = customer.invoices.last()
        invoice_pdf_url = invoice.pdf.pdf_file.url if invoice and invoice.pdf else None
        customer_invoices.append({
            'customer': customer,
            'invoice_pdf_url': invoice_pdf_url
        })

    paginator = Paginator(customer_invoices, 10)  
    try:
        customer_invoices = paginator.page(page)
    except PageNotAnInteger:
        customer_invoices = paginator.page(1)
    except EmptyPage:
        customer_invoices = paginator.page(paginator.num_pages)
        
    context = {
        'team_leaders': UserProfile.objects.filter(role='Team Leader').all(),
        'payment_status_choices': PaidCustomer.PAYMENT_STATUS_CHOICES,
        'payment_methods_choices': PaymentMethod.objects.all(),
        'customer_invoices': customer_invoices,
        'is_superuser': user.is_superuser,
        'search_query': search_query,
        'disposition': disposition,
        'sub_disposition': sub_disposition,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'request': request,
    }

    return render(request, 'paid_customers.html', context)

@transaction.atomic
def verify_customer(request):
    if request.method == 'GET' and 'customer_id' in request.GET:
        try:
            customer_id = request.GET['customer_id']
            paid_customer = PaidCustomer.objects.select_for_update().get(id=customer_id)
            paid_customer.verified = True
            paid_customer.save()

            amount_in_words = num2words(paid_customer.amount_paid, to='currency', lang='en_IN')
            amount_in_words = amount_in_words.replace('euro', 'rupees').replace('cents', '').strip()
            if amount_in_words.endswith('zero'):
                amount_in_words = amount_in_words.rsplit(' ', 1)[0].rstrip(', ').strip()
            amount_in_words = re.split(r'rupees', amount_in_words, flags=re.IGNORECASE)[0] + "RUPEES"
            amount_in_words = amount_in_words.upper()

            company = Company.objects.first()
            if not company:
                raise ObjectDoesNotExist("No company found. Please create a company object in the database.")

            invoice = Invoice.objects.create(
                unique_invoice_number=str(uuid.uuid4().int)[:8],
                customer=paid_customer,
                amount_in_words=amount_in_words,
                company=company,
                date=timezone.localtime(timezone.now())
            )

            html_template = 'print_invoice_template.html'
            context = {
                'invoice': invoice,
                'company': company,
                'paid_customer': paid_customer,
                'static_url': settings.STATIC_URL, 
                'media_url': settings.MEDIA_ROOT,  
            }

            html_string = render_to_string(html_template, context, request=request)
            css_path = os.path.join(settings.STATICFILES_DIRS[0], 'css', 'print_invoice_template.css')
            options = {
                'page-size': 'A4',
                'encoding': 'UTF-8',
                'no-outline': None,
                'enable-local-file-access': None,
                'quiet': '',
            }

            config = pdfkit.configuration(wkhtmltopdf=r"wkhtmltopdf.exe")
            pdf_file = pdfkit.from_string(html_string, False, options=options, configuration=config, css=css_path)
            pdf_filename = f'{paid_customer.lead.full_name}-{paid_customer.lead.contact_number}-Invoice.pdf'
            invoice_pdf = InvoicePDF.objects.create(
                invoice_object=invoice
            )
            invoice_pdf.pdf_file.save(pdf_filename, BytesIO(pdf_file))
            invoice.pdf = invoice_pdf
            invoice.save()

            messages.success(request, 'Customer verified successfully and invoice created.')
            return redirect('paid_customers')

        except ObjectDoesNotExist as e:
            messages.error(request, str(e))
            return redirect('paid_customers')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('paid_customers')

    messages.error(request, 'Invalid request.')
    return redirect('paid_customers')

@login_required
def create_paid_customer(request):
    if request.method == "POST":
        form = PaidCustomerForm(request.POST, request.FILES)
        if form.is_valid():
            paid_customer = form.save(commit=False)
            try:
                lead = Lead.objects.get(contact_number=paid_customer.contact_number)
                paid_customer.lead = lead
            except Lead.DoesNotExist:
                paid_customer.lead = None
            
            existing_paid_customer = PaidCustomer.objects.filter(contact_number=paid_customer.contact_number).first()
            if existing_paid_customer:
                paid_customer.customer_id = existing_paid_customer.customer_id
            else:
                paid_customer.customer_id = str(uuid4().int)[:12]
            
            paid_customer.save()
            return redirect('paid_customers')
    else:
        form = PaidCustomerForm()

    return render(request, 'create_paid_customer.html', {'form': form})

@login_required
def autocomplete_leads(request):
    query = request.GET.get('query')
    field = request.GET.get('field')
    user = request.user

    if user.is_superuser:
        leads = Lead.objects.filter(
            Q(full_name__icontains=query) | Q(contact_number__icontains=query)
        ).values('full_name', 'contact_number')
    elif user.role == 'Team Leader':
        team = Team.objects.filter(leader=user)
        leads = Lead.objects.filter(
            Q(full_name__icontains=query) | Q(contact_number__icontains=query),
            assigned_to_team=team
        ).values('full_name', 'contact_number')
    elif user.role == 'Agent':
        leads = Lead.objects.filter(
            Q(full_name__icontains=query) | Q(contact_number__icontains=query),
            assigned_to=user
        ).values('full_name', 'contact_number')
    else:
        leads = []

    return JsonResponse(list(leads), safe=False)




@login_required
def edit_paid_customer(request, customerId):
    user = request.user
    customer = get_object_or_404(PaidCustomer, id=customerId)

    # Check if the customer is verified
    if customer.verified:
        # Only superuser can edit a verified customer
        if not user.is_superuser:
            return redirect('paid_customers')
    else:
        # If not verified, check for team leader and agent roles
        if not user.is_superuser:
            if user.profile.role == 'Team Leader':
                team = Team.objects.filter(leader=user.profile).first()
                if team:
                    team_members = team.agents.all()
                    if not customer.lead or customer.lead.assigned_to not in team_members:
                        return redirect('paid_customers')
                else:
                    return redirect('paid_customers')
            elif user.profile.role == 'Agent':
                if not customer.lead or customer.lead.assigned_to != user.profile:
                    return redirect('paid_customers')
            else:
                return redirect('paid_customers')

    if request.method == "POST":
        form = PaidCustomerForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            paid_customer = form.save(commit=False)
            try:
                lead = Lead.objects.get(contact_number=paid_customer.contact_number)
                paid_customer.lead = lead
            except Lead.DoesNotExist:
                paid_customer.lead = None
            
            paid_customer.save()
            return redirect('paid_customers')
    else:
        form = PaidCustomerForm(instance=customer)
    return render(request, 'edit_paid_customer.html', {'form': form})




def delete_paid_customer(request, customerId):
    customer = get_object_or_404(PaidCustomer, id=customerId)
    if request.method == "POST":
        customer.delete()
        return redirect('paid_customers')
    return render(request, 'delete_paid_customer.html', {'customer': customer})

@login_required
def export_paid_customers(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="paid_customers.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Customer ID', 'Contact Number', 'Full Name', 'Payment Date',
        'Package Name', 'Amount Paid', 'Amount with Gst', 'Transaction ID',
        'Payment Status', 'Payment Method', 'PAN Number', 'Agent Name', 'TL Name'
    ])

    paid_customers = PaidCustomer.objects.all().select_related(
        'lead__assigned_to', 'lead__assigned_to_team'
    )

    for paid_customer in paid_customers:
        agent_name = paid_customer.lead.assigned_to.user.get_full_name() if paid_customer.lead.assigned_to else ''
        tl_name = paid_customer.lead.assigned_to_team.leader.user.get_full_name() if paid_customer.lead.assigned_to_team else ''
        
        writer.writerow([
            paid_customer.date,
            paid_customer.customer_id,
            paid_customer.contact_number,
            paid_customer.lead.full_name if paid_customer.lead else '',
            paid_customer.payment_date,
            paid_customer.package.name,
            paid_customer.amount_paid,
            paid_customer.amount_with_gst,  
            paid_customer.transaction_id,
            paid_customer.payment_status,
            paid_customer.payment_method.name,
            paid_customer.pan_number,
            agent_name,
            tl_name,
        ])

    return response

@login_required
def create_or_update_company(request):
    company_instance = Company.objects.first()

    if request.method == 'POST':
        if company_instance:
            form = CompanyForm(request.POST, request.FILES, instance=company_instance)
        else:
            form = CompanyForm(request.POST, request.FILES)

        if form.is_valid():
            try:
                form.save()
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = CompanyForm(instance=company_instance)

    context = {
        'form': form,
        'company': company_instance,
    }
    return render(request, 'invoice_template.html', context)

###################################################################################################################################################

@login_required
def complaints_list(request):
    complaints = Complaint.objects.filter(user=request.user.profile)

    search_query = request.GET.get('search_query', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    sort = request.GET.get('sort', 'id')

    if start_date:
        complaints = complaints.filter(created_at__date__gte=start_date)
    if end_date:
        complaints = complaints.filter(created_at__date__lte=end_date)
    if status:
        complaints = complaints.filter(status=status)
    if priority:
        complaints = complaints.filter(priority=priority)
    if search_query:
        complaints = complaints.filter(subject__icontains=search_query)
    if sort:
        complaints = complaints.order_by(sort)
    else:
        complaints = complaints.order_by('-id')

    # Pagination
    paginator = Paginator(complaints, 10) 

    page_number = request.GET.get('page')
    try:
        complaints = paginator.page(page_number)
    except PageNotAnInteger:
        complaints = paginator.page(1)
    except EmptyPage:
        complaints = paginator.page(paginator.num_pages)

    context = {
        'complaints': complaints,
        'complaint_status_choices': Complaint.STATUS_CHOICES,
        'complaint_priority_choices': Complaint.PRIORITY_CHOICES,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
        'status': status,
        'priority': priority,
        'sort': sort,
    }

    return render(request, 'complaints_list.html', context)


@login_required
def create_complaint(request):
    if request.method == 'POST':
        form = ComplaintForm(request.POST)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.user = request.user.profile 
            complaint.save()
            return redirect('complaints_list')
    else:
        form = ComplaintForm()
    
    return render(request, 'create_complaint.html', {'form': form})

@login_required
def edit_complaint(request, complaint_id):
    complaint = get_object_or_404(Complaint, id=complaint_id)

    if request.method == 'POST':
        form = ComplaintForm(request.POST, instance=complaint)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.user = request.user.profile 
            complaint.save()
            return redirect('complaints_list')
    else:
        form = ComplaintForm(instance=complaint)

    context = {
        'form': form,
    }

    return render(request, 'edit_complaint.html', context)

@login_required
def delete_complaint(request, complaint_id):
    complaint = get_object_or_404(Complaint, id=complaint_id)
    if request.method == 'POST':
        complaint.delete()
        messages.success(request, 'Complaint deleted successfully.')
        return redirect('complaints_list')
    return render(request, 'delete_complaint.html', {'complaint': complaint})

#####################################################################################################################################################


@login_required
def attendance(request):
    user = request.user

    if user.is_superuser:
        attendances = Attendance.objects.exclude(user=user.profile)
    elif user.profile.role == 'Team Leader':
        team = user.profile.teams_as_leader.first()
        agents = team.agents.exclude(id=user.profile.id)
        attendances = Attendance.objects.filter(user__in=agents).select_related('user')
    elif user.profile.role == 'Agent':
        attendances = Attendance.objects.filter(user=user.profile).select_related('user')
    else:
        attendances = Attendance.objects.none()

    search_query = request.GET.get('search_query', '')
    if search_query:
        attendances = attendances.filter(
            Q(user__user__first_name__icontains=search_query) |
            Q(user__user__last_name__icontains=search_query) |
            Q(user__user__email__icontains=search_query)
        )

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        attendances = attendances.filter(date__gte=start_date)
    if end_date:
        attendances = attendances.filter(date__lte=end_date)

    local_now = timezone.localtime(timezone.now())
    today = local_now.date()
    date_filter = request.GET.get('date_filter')
    if date_filter == 'today':
        attendances = attendances.filter(date=today)
    elif date_filter == 'this_week':
        start_of_week = today - timezone.timedelta(days=today.weekday())
        attendances = attendances.filter(date__gte=start_of_week, date__lte=today)
    elif date_filter == 'this_month':
        start_of_month = today.replace(day=1)
        attendances = attendances.filter(date__gte=start_of_month, date__lte=today)
    elif date_filter == 'last_6_months':
        six_months_ago = today - timezone.timedelta(days=6*30)
        attendances = attendances.filter(date__gte=six_months_ago, date__lte=today)
    elif date_filter == 'this_year':
        start_of_year = today.replace(month=1, day=1)
        attendances = attendances.filter(date__gte=start_of_year, date__lte=today)

    role = request.GET.get('role')
    if role:
        attendances = attendances.filter(user__role=role)

    status = request.GET.get('status')
    if status:
        attendances = attendances.filter(status=status)

    on_time_late = request.GET.get('on_time_late')
    if on_time_late:
        attendances = attendances.filter(on_time_late=on_time_late)

    attendances = attendances.prefetch_related('breaks__break_type')
    attendances = attendances.order_by('-id')

    # Pagination
    paginator = Paginator(attendances, 10)  

    page_number = request.GET.get('page')
    try:
        attendances = paginator.page(page_number)
    except PageNotAnInteger:
        attendances = paginator.page(1)
    except EmptyPage:
        attendances = paginator.page(paginator.num_pages)

    context = {
        'attendances': attendances,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
        'role': role,
        'on_time_late': on_time_late,
        'status': status,
        'date_filter': date_filter,
    }

    return render(request, 'attendance.html', context)



def update_regulation_reason(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)

    if request.method == 'POST':
        new_regulation_reason = request.POST.get('regulation_reason')
        attendance.regulation_reason = new_regulation_reason
        attendance.save()
        messages.success(request, 'Regulation reason updated successfully.')

        return redirect('attendance')

    return render(request, 'attendance.html', {'attendance': attendance})



@login_required
def export_attendance(request):
    export_type = request.GET.get('export_type', '')

    if export_type == 'team_leaders':
        attendances = Attendance.objects.filter(user__role='Team Leader')
        headers = [
            'Date & Day', 'TL Name', 'Login Time', 'Logout Time', 'Status', 
            'On Time or Late', 'Total Login Time', 'Total Break Time', 'Regulation', 'TEA', 'LUNCH', 
            'TL BRIEFING', 'QUALITY BRIEFING', 'FLOOR MEETING'
        ]
    elif export_type == 'agents':
        attendances = Attendance.objects.filter(user__role='Agent')
        headers = [
            'Date & Day', 'Agent Name', 'TL Name', 'Login Time', 'Logout Time', 'Status', 
            'On Time or Late', 'Total Login Time', 'Total Break Time', 'Regulation', 'TEA', 'LUNCH', 
            'TL BRIEFING', 'QUALITY BRIEFING', 'FLOOR MEETING'
        ]
    else:
        attendances = Attendance.objects.all()
        headers = [
            'Date & Day', 'Agent Name', 'TL Name', 'Login Time', 'Logout Time', 'Status', 
            'On Time or Late', 'Total Login Time', 'Total Break Time', 'Regulation', 'TEA', 'LUNCH', 
            'TL BRIEFING', 'QUALITY BRIEFING', 'FLOOR MEETING'
        ]

    attendances = attendances.select_related('user').prefetch_related('breaks')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.append(headers)

    for attendance in attendances:
        tea_time = 0
        lunch_time = 0
        tl_briefing_time = 0
        quality_briefing_time = 0
        floor_meeting_time = 0

        for break_obj in attendance.breaks.all():
            if break_obj.break_type.name == 'TEA':
                tea_time += break_obj.break_duration()
            elif break_obj.break_type.name == 'LUNCH':
                lunch_time += break_obj.break_duration()
            elif break_obj.break_type.name == 'TL BRIEFING':
                tl_briefing_time += break_obj.break_duration()
            elif break_obj.break_type.name == 'QUALITY BRIEFING':
                quality_briefing_time += break_obj.break_duration()
            elif break_obj.break_type.name == 'FLOOR MEETING':
                floor_meeting_time += break_obj.break_duration()

        tea_time_str = f"{tea_time // 60} hours {tea_time % 60} minutes"
        lunch_time_str = f"{lunch_time // 60} hours {lunch_time % 60} minutes"
        tl_briefing_time_str = f"{tl_briefing_time // 60} hours {tl_briefing_time % 60} minutes"
        quality_briefing_time_str = f"{quality_briefing_time // 60} hours {quality_briefing_time % 60} minutes"
        floor_meeting_time_str = f"{floor_meeting_time // 60} hours {floor_meeting_time % 60} minutes"

        if attendance.user.role == 'Agent':
            team_leader_name = 'N/A'
            team = attendance.user.teams_as_agent.first()
            if team:
                team_leader_name = team.leader.user.get_full_name()
            row = [
                f"{attendance.date} ({attendance.day})",
                attendance.user.user.get_full_name(), 
                team_leader_name,
                attendance.login_time.strftime('%H:%M') if attendance.login_time else 'N/A',
                attendance.logout_time.strftime('%H:%M') if attendance.logout_time else 'N/A',
                attendance.status,
                attendance.on_time_late,
                f"{attendance.total_login_time_hours} hours {attendance.total_login_time_minutes} minutes",
                f"{attendance.total_break_time_hours} hours {attendance.total_break_time_minutes} minutes",
                attendance.regulation_reason,
                tea_time_str,
                lunch_time_str,
                tl_briefing_time_str,
                quality_briefing_time_str,
                floor_meeting_time_str
            ]
        else:
            row = [
                f"{attendance.date} ({attendance.day})",
                attendance.user.user.get_full_name(),
                attendance.login_time.strftime('%H:%M') if attendance.login_time else 'N/A',
                attendance.logout_time.strftime('%H:%M') if attendance.logout_time else 'N/A',
                attendance.status,
                attendance.on_time_late,
                f"{attendance.total_login_time_hours} hours {attendance.total_login_time_minutes} minutes",
                f"{attendance.total_break_time_hours} hours {attendance.total_break_time_minutes} minutes",
                attendance.regulation_reason,
                tea_time_str,
                lunch_time_str,
                tl_briefing_time_str,
                quality_briefing_time_str,
                floor_meeting_time_str
            ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    wb.save(response)

    return response


####################################################################################################################################################

def sales(request):
    user = request.user
    search_query = request.GET.get('search_query', '')
    team_leader_id = request.GET.get('team_leader', '')
    sales_achievement = request.GET.get('sales_achievement', '')

    if user.is_superuser:
        agents = UserProfile.objects.filter(role='Agent')
    elif user.profile.role == 'Team Leader':
        team = Team.objects.filter(leader=user.profile).first()
        if team:
            agents = team.agents.all()
        else:
            agents = UserProfile.objects.none()
    elif user.profile.role == 'Agent':
        teams = user.profile.teams_as_agent.all()
        agents = UserProfile.objects.filter(teams_as_agent__in=teams, role='Agent').distinct()
    else:
        agents = UserProfile.objects.none()

    if search_query:
        agents = agents.filter(Q(user__first_name__icontains=search_query) | Q(user__last_name__icontains=search_query))

    if team_leader_id:
        try:
            team = Team.objects.get(leader__user__id=team_leader_id)
            agents = agents.filter(teams_as_agent=team)
        except Team.DoesNotExist:
            team_leader = get_object_or_404(UserProfile, user__id=team_leader_id, role='Team Leader')
            messages.error(request, f"No team exists for Team Leader: {team_leader.user.get_full_name()}")

    sales_summary = []

    for agent in agents:
        total_sales = Invoice.objects.filter(customer__lead__assigned_to=agent, customer__payment_status='completed').aggregate(total_sales=Sum('customer__amount_with_gst'))['total_sales'] or 0
        number_of_customers = PaidCustomer.objects.filter(lead__assigned_to=agent, payment_status='completed').count()
        achievement_percentage = (total_sales / (agent.commitment or 1)) * 100
        assigned_leads = Lead.objects.filter(assigned_to=agent).count()
        conversion_rate = (number_of_customers / assigned_leads) * 100 if assigned_leads else 0
        total_invoice_generated = Invoice.objects.filter(customer__lead__assigned_to=agent, customer__payment_status='completed').count()
        attendances = Attendance.objects.filter(user=agent.user.id)
        total_days = attendances.count()
        if total_days == 0:
            attendance_percentage = 0.0
        else:
            present_days = attendances.filter(status='Present').count()
            half_day_days = attendances.filter(status='Half day').count()
            attendance_percentage = ((present_days + half_day_days / 2) / total_days) * 100
        
        team_leader = agent.teams_as_agent.first().leader.user.get_full_name() if agent.teams_as_agent.exists() else 'N/A'
        
        sales_summary.append({
            'agent': agent,
            'team_leader': team_leader,
            'attendance': round(attendance_percentage, 2),
            'lead_count': assigned_leads,
            'conversion': round(conversion_rate, 2),
            'total_invoice_generated':total_invoice_generated,
            'sales': round(total_sales, 2),
            'achievements': round(achievement_percentage, 2),
            'commitment': agent.commitment,
        })

    sales_summary.sort(key=lambda x: x['achievements'], reverse=True)
    for idx, record in enumerate(sales_summary):
        record['rank'] = idx + 1

    if sales_achievement:
        sales_summary = [s for s in sales_summary if s['achievements'] >= int(sales_achievement)]

    context = {
        'sales_summary': sales_summary,
        'search_query': search_query,
        'team_leader_id': team_leader_id,
        'sales_achievement': sales_achievement,
        'team_leaders': UserProfile.objects.filter(role='Team Leader'),
    }

    return render(request, 'sales.html', context)

def export_sales(request):
    agents = UserProfile.objects.filter(role='Agent')
    sales_summary = []

    for agent in agents:
        total_sales = Invoice.objects.filter(customer__lead__assigned_to=agent, customer__payment_status='completed').aggregate(total_sales=Sum('amount_with_gst'))['total_sales'] or 0
        number_of_customers = PaidCustomer.objects.filter(lead__assigned_to=agent, payment_status='completed').count()
        achievement_percentage = (total_sales / (agent.commitment or 1)) * 100
        assigned_leads = Lead.objects.filter(assigned_to=agent).count()
        conversion_rate = (number_of_customers / assigned_leads) * 100 if assigned_leads else 0
        attendances = Attendance.objects.filter(user=agent.user.id)
        total_days = attendances.count()
        attendance_percentage = ((attendances.filter(status='Present').count() + attendances.filter(status='Half day').count() / 2) / total_days) * 100 if total_days else 0
        team_leader = agent.teams_as_agent.first().leader.user.get_full_name() if agent.teams_as_agent.exists() else 'N/A'
        sales_summary.append({
            'agent': agent,
            'team_leader': team_leader,
            'attendance': round(attendance_percentage, 2),
            'lead_count': assigned_leads,
            'conversion': round(conversion_rate, 2),
            'sales': round(total_sales, 2),
            'achievements': round(achievement_percentage, 2),
            'commitment': agent.commitment,
        })

    sales_summary.sort(key=lambda x: x['achievements'], reverse=True)
    for idx, record in enumerate(sales_summary):
        record['rank'] = idx + 1

    if not sales_summary:
        messages.error(request, "No sales data meets the specified achievement criteria.")
        return HttpResponseRedirect(reverse('sales'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Summary"
    headers = ['Rank', 'Agent', 'Team Leader', 'Attendance', 'Lead Count', 'Conversion', 'Sales', 'Achievements', 'Commitment']
    ws.append(headers)

    for record in sales_summary:
        ws.append([
            record['rank'],
            record['agent'].user.get_full_name(),
            record['team_leader'],
            f"{floatformat(record['attendance'], 2)}%",
            record['lead_count'],
            f"{floatformat(record['conversion'], 2)}%",
            round(record['sales'], 2),
            f"{floatformat(record['achievements'], 2)}%",
            round(record['commitment'], 2),
        ])

    currency_format = '_("₹"* #,##,##0.00_);_("₹"* (#,##,##0.00);_("₹"* "-"??_);_(@_)'
    for col_num, header in enumerate(headers, start=1):
        if header in ['Sales', 'Commitment']:
            for cell in ws[openpyxl.utils.get_column_letter(col_num)]:
                cell.number_format = currency_format

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_summary.xlsx"'
    wb.save(response)

    messages.success(request, "Sales data exported successfully.")
    return response

def agent_sales_history(request, agent_id):
    agent = get_object_or_404(User, id=agent_id)
    agent_profile = get_object_or_404(UserProfile, user=agent)

    if request.method == 'POST':
        form = UpdateSalesForm(request.POST)
        if form.is_valid():
            commitment = form.cleaned_data['commitment']
            agent_profile.commitment = commitment
            agent_profile.save()
            
            record_agent_sales_history(agent=agent, commitment=commitment, updated_by=request.user)
            return redirect('agent_sales_history', agent_id=agent.id)
    else:
        initial_data = {'commitment': agent_profile.commitment} if agent_profile.commitment is not None else {}
        form = UpdateSalesForm(initial=initial_data)
            
    sales_history = AgentSalesHistory.objects.filter(agent=agent).order_by('-date_time')
    
    context = {
        'agent': agent,
        'form': form,
        'sales_history': sales_history,
    }
    return render(request, 'agent_sales_history.html', context)

#####################################################################################################################################################

@login_required
def settings_view(request):
    return render(request, 'settings.html')

@login_required
def other_administrative_settings(request):
    if request.method == 'POST':
        if 'break_type_submit' in request.POST:
            break_type_form = BreakTypeForm(request.POST)
            package_form = PackageForm()
            sub_disposition_form = SubDispositionForm()
            if break_type_form.is_valid():
                break_type_form.save()
                return redirect('other_administrative_settings')
        elif 'package_submit' in request.POST:
            break_type_form = BreakTypeForm()
            package_form = PackageForm(request.POST)
            sub_disposition_form = SubDispositionForm()
            if package_form.is_valid():
                package_form.save()
                return redirect('other_administrative_settings')
        elif 'sub_disposition_submit' in request.POST:
            break_type_form = BreakTypeForm()
            package_form = PackageForm()
            sub_disposition_form = SubDispositionForm(request.POST)
            if sub_disposition_form.is_valid():
                sub_disposition_form.save()
                return redirect('other_administrative_settings')
        elif 'payment_method_submit' in request.POST:
                payment_method_form = PaymentMethodForm(request.POST)
                if payment_method_form.is_valid():
                    payment_method_form.save()
                    return redirect('other_administrative_settings')
    else:
        break_type_form = BreakTypeForm()
        package_form = PackageForm()
        sub_disposition_form = SubDispositionForm()
        payment_method_form = PaymentMethodForm()

    all_breaks = BreakType.objects.all()
    all_package = Package.objects.all()
    all_sub_disposition = SubDisposition.objects.all()
    all_payment_method = PaymentMethod.objects.all()

    context = {
        'all_breaks':all_breaks,
        'all_package':all_package,
        'all_sub_disposition':all_sub_disposition,
        'all_payment_method':all_payment_method,
        'break_type_form': break_type_form,
        'package_form': package_form,
        'sub_disposition_form': sub_disposition_form,
        'payment_method_form': payment_method_form
    }

    return render(request, 'other_administrative_settings.html', context)

@login_required
def delete_break(request, breaktype_id):
    break_instance = get_object_or_404(BreakType, id=breaktype_id)
    if request.method == 'POST':
        break_instance.delete()
        messages.success(request, 'Break deleted successfully.')
    return redirect(reverse('other_administrative_settings')) 

@login_required
def delete_package(request, package_id):
    package_instance = get_object_or_404(Package, id=package_id)
    if request.method == 'POST':
        package_instance.delete()
        messages.success(request, 'Package deleted successfully.')
    return redirect(reverse('other_administrative_settings')) 

@login_required
def delete_sub_disposition(request, sub_disposition_id):
    sub_disposition_instance = get_object_or_404(SubDisposition, id=sub_disposition_id)
    if request.method == 'POST':
        sub_disposition_instance.delete()
        messages.success(request, 'Sub Disposition deleted successfully.')
    return redirect(reverse('other_administrative_settings')) 

@login_required
def delete_payment_method(request, payment_method_id):
    payment_method_instance = get_object_or_404(PaymentMethod, id=payment_method_id)
    if request.method == 'POST':
        payment_method_instance.delete()
        messages.success(request, 'Payment Method deleted successfully.')
    return redirect(reverse('other_administrative_settings')) 


#####################################################################################################################################################

@login_required
def analytics(request):
    user = request.user
    
    if user.is_superuser:
        total_leads = Lead.objects.all().count()
        total_paid_customers = PaidCustomer.objects.all().count()
        total_invoice_generated = Invoice.objects.all().count()
        invoices = Invoice.objects.all()
        total_revenue = sum(invoice.customer.amount_with_gst for invoice in invoices)
        total_complaints = Complaint.objects.all().count()
        total_attendances = Attendance.objects.count()
        total_present = Attendance.objects.filter(status='Present').count()
        total_absent = Attendance.objects.filter(status='Absent').count()
        total_half_day = Attendance.objects.filter(status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100
        else:
            attendance_rate = 0

        context = {
            'total_leads': total_leads,
            'total_invoice_generated': total_invoice_generated,
            'total_complaints': total_complaints,
            'attendance_rate': attendance_rate,
            'total_paid_customers': total_paid_customers,
            'total_revenue': total_revenue,
            'total_present': total_present,
            'total_absent': total_absent,
            'total_half_day': total_half_day
        }

    elif user.profile.role == 'Team Leader':

        team = user.profile.teams_as_leader.first() 
        if team:
            agents = team.agents.filter(role='Agent')  
        else:
            agents = UserProfile.objects.none() 
        
        total_leads = Lead.objects.filter(assigned_to__in=agents).count()
        total_paid_customers = PaidCustomer.objects.filter(lead__assigned_to__in=agents).count()
        total_invoice_generated = Invoice.objects.filter(customer__lead__assigned_to__in=agents).count()
        invoices = Invoice.objects.filter(customer__lead__assigned_to__in=agents)
        total_revenue = sum(invoice.customer.amount_with_gst for invoice in invoices)
        total_complaints = Complaint.objects.filter(user__in=agents).count()
        total_attendances = Attendance.objects.filter(user__in=agents).count()
        total_present = Attendance.objects.filter(user__in=agents, status='Present').count()
        total_absent = Attendance.objects.filter(user__in=agents, status='Absent').count()
        total_half_day = Attendance.objects.filter(user__in=agents, status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100
        else:
            attendance_rate = 0

        context = {
            'total_leads': total_leads,
            'total_invoice_generated': total_invoice_generated,
            'total_complaints': total_complaints,
            'attendance_rate': attendance_rate,
            'total_paid_customers': total_paid_customers,
            'total_revenue': total_revenue,
            'total_present': total_present,
            'total_absent': total_absent,
            'total_half_day': total_half_day
        }

    elif user.profile.role == 'Agent':
        total_leads = Lead.objects.filter(assigned_to=user.profile).count()
        total_paid_customers = PaidCustomer.objects.filter(lead__assigned_to=user.profile).count()
        total_invoice_generated = Invoice.objects.filter(customer__lead__assigned_to=user.profile).count()
        invoices = Invoice.objects.filter(customer__lead__assigned_to=user.profile)
        total_revenue = sum(invoice.customer.amount_with_gst for invoice in invoices)
        total_complaints = Complaint.objects.filter(user=user.profile).count()
        total_attendances = Attendance.objects.filter(user=user.profile).count()
        total_present = Attendance.objects.filter(user=user.profile, status='Present').count()
        total_absent = Attendance.objects.filter(user=user.profile, status='Absent').count()
        total_half_day = Attendance.objects.filter(user=user.profile, status='Half day').count()

        if total_attendances > 0:
            attendance_rate = ((total_present + total_half_day) / total_attendances) * 100
        else:
            attendance_rate = 0

        context = {
            'total_leads': total_leads,
            'total_invoice_generated': total_invoice_generated,
            'total_complaints': total_complaints,
            'attendance_rate': attendance_rate,
            'total_paid_customers': total_paid_customers,
            'total_revenue': total_revenue,
            'total_present': total_present,
            'total_absent': total_absent,
            'total_half_day': total_half_day
        }

    return render(request, 'analytics.html', context)

#####################################################################################################################################################

@login_required
def reports(request):
    user = request.user
    team_leader_id = request.GET.get('team_leader', 'all_teams')
    agent_id = request.GET.get('agent')

    leads = None
    sub_dispositions = None

    if user.profile.role == 'Team Leader':
        team_leaders = UserProfile.objects.filter(id=user.profile.id)
        agents = UserProfile.objects.filter(role='Agent', teams_as_agent__leader=user.profile).order_by('user__first_name', 'user__last_name')
    else:
        team_leaders = UserProfile.objects.filter(role='Team Leader').order_by('user__first_name', 'user__last_name')
        if team_leader_id and team_leader_id != 'all_teams':
            agents = UserProfile.objects.filter(role='Agent', teams_as_agent__leader_id=team_leader_id).order_by('user__first_name', 'user__last_name')
        else:
            agents = UserProfile.objects.filter(role='Agent').order_by('user__first_name', 'user__last_name')

    if agent_id:
        leads = Lead.objects.filter(assigned_to_id=agent_id).values('assigned_to_id').annotate(count=Count('id'))
        sub_dispositions = Lead.objects.filter(assigned_to_id=agent_id).values('sub_disposition__name').annotate(lead_count=Count('id'))

    context = {
        'agents': agents,
        'leads': leads,
        'sub_dispositions': sub_dispositions,
        'team_leaders': team_leaders,
        'current_team_leader': team_leader_id,
        'current_agent': agent_id,
    }

    return render(request, 'reports.html', context)

def get_leads_by_sub_disposition(request):
    sub_disposition_name = request.GET.get('sub_disposition')
    agent_id = request.GET.get('agent_id')

    if sub_disposition_name:
        leads = Lead.objects.filter(sub_disposition__name=sub_disposition_name, assigned_to=agent_id)

        leads = leads.values(
            'id', 'date', 'full_name', 'contact_number', 'state', 'capital',
            'assigned_to__user__first_name', 'assigned_to__user__last_name', 'assigned_to__user__username',
            'assigned_to__teams_as_agent__name', 'assigned_to__teams_as_agent__leader__user__first_name', 'assigned_to__teams_as_agent__leader__user__last_name',
            'disposition', 'remark', 'sub_disposition__name'
        )
        
        leads_list = list(leads)
        for lead in leads_list:
            lead['assigned_to_full_name'] = f"{lead['assigned_to__user__first_name']} {lead['assigned_to__user__last_name']}".strip()
            lead['assigned_to_team_leader_full_name'] = f"{lead['assigned_to__teams_as_agent__leader__user__first_name']} {lead['assigned_to__teams_as_agent__leader__user__last_name']}".strip()
            del lead['assigned_to__user__first_name']
            del lead['assigned_to__user__last_name']
            del lead['assigned_to__teams_as_agent__leader__user__first_name']
            del lead['assigned_to__teams_as_agent__leader__user__last_name']

        return JsonResponse(leads_list, safe=False)
    
    return JsonResponse({'error': 'Sub disposition not provided'}, status=400)

@login_required
def export_lead_report(request):
    user_id = request.GET.get('user_id')
    if not user_id:
        return HttpResponse(status=400) 
    
    try:
        user = UserProfile.objects.get(user__id=user_id)
    except UserProfile.DoesNotExist:
        return HttpResponse(status=404)  
    
    if user.role == 'Team Leader':
        team_members = UserProfile.objects.filter(teams_as_agent__leader=user)
        leads = Lead.objects.filter(assigned_to__in=team_members)
    else:
        leads = Lead.objects.filter(assigned_to=user)

    sub_dispositions = SubDisposition.objects.values_list('name', flat=True).distinct()

    disposition_summary = {}
    for sub_disposition in sub_dispositions:
        count = leads.filter(sub_disposition__name=sub_disposition).count()
        disposition_summary[sub_disposition] = count
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="lead_report_user_{user_id}.csv"'
    writer = csv.writer(response)

    summary_headers = list(disposition_summary.keys())
    writer.writerow(summary_headers)

    summary_data = [disposition_summary.get(header, 0) for header in summary_headers]
    writer.writerow(summary_data)
    writer.writerow([])
    
    writer.writerow(['Date', 'Agent Name', 'TL Name', 'Customer Name', 'Contact Number', 'State', 
                     'Capital', 'Disposition', 'Sub Disposition', 'Remark', 'Reminder'])
    
    for lead in leads:
        writer.writerow([
            lead.date,
            lead.assigned_to.user.get_full_name(),
            lead.assigned_to.teams_as_agent.first().leader.user.get_full_name() if lead.assigned_to.teams_as_agent.exists() else '',
            lead.full_name,
            lead.contact_number,
            lead.state,
            lead.capital,
            lead.disposition,
            lead.sub_disposition.name if lead.sub_disposition else '',
            lead.remark,
            lead.reminder
        ])
    
    return response

